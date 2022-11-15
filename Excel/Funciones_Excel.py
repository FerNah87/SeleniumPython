import openpyxl

class FunExcel():
    def __init__(self, driver):
        self.driver = driver
    #1ra funcion cuenta las filas
    def getRowCount(file, path, sheetName):
        Workbook = openpyxl.load_workbook(path)
        sheet = Workbook[sheetName]
        return (sheet.max_row)
    #2da funcion cuenta las columnas
    def getColumncount(file, sheetName):
        Workbook = openpyxl.load_workbook(file)
        sheet = Workbook[sheetName]
        return (sheet.max_column)
    #3ra funcion lee la hoja de excel
    def readData(file, path, sheetName, rownum, colnum):
        Workbook = openpyxl.load_workbook(path)
        sheet = Workbook[sheetName]
        return sheet.cell(row = rownum, column = colnum).value
    #4ta funcion escribir en el excel
    def writeData(file, path, sheetName, rownum, colnum, data):
        Workbook = openpyxl.load_workbook(path)
        sheet = Workbook[sheetName]
        sheet.cell(row=rownum, column=colnum).value = data
        Workbook.save(path)